import json
import os
import logging
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

from backend.config import settings
from backend.database import get_all_appointments, get_patient_documents, init_db

logger = logging.getLogger(__name__)


BOOKING_COLUMNS = SHEET_COLUMNS = [
    "Appointment ID",
    "Patient Name",
    "Phone",
    "Age",
    "Location",
    "Date",
    "Time",
    "Reason",
    "Consultation Type",
    "Status",
    "Payment Status",
    "ID Card",
    "ID Card Image URL",
    "Payment Screenshot URL",
    "Reports/Prescriptions URLs",
    "Reports Data",
    "Created At",
    "Updated At",
    "Source",
    "Reminder Sent",
    "Followup Sent",
    "Followup Response",
]


def _stringify(value) -> str:
    if value is None:
        return ""
    return str(value)


def appointment_rows() -> list[list[str]]:
    """Build rows matching SHEET_COLUMNS order for Excel export."""
    init_db()
    rows = []
    for appt in get_all_appointments():
        # Parse details_json for contact number
        contact_number = ""
        details_json = appt.get("details_json")
        if details_json:
            try:
                parsed = json.loads(details_json)
                contact_number = parsed.get("contact_number", "")
            except Exception:
                pass
        display_phone = contact_number if contact_number else _stringify(appt.get("phone"))

        # ID card display
        id_card_val = _stringify(appt.get("id_card"))
        id_card_image = appt.get("id_card_image_path")
        id_card_display = "ID image on file" if id_card_image else id_card_val

        # Status mapping
        status_map = {
            "booked": "Created", "confirmed": "Confirmed", "cancelled": "Cancelled",
            "rescheduled": "Rescheduled", "no_show": "No Show", "checked_in": "Checked In",
        }
        raw_status = _stringify(appt.get("status"))
        calendar_status = status_map.get(raw_status, raw_status)

        # Reports data summary
        reports_data = _stringify(appt.get("reports_data_json", ""))
        if reports_data:
            try:
                reports_parsed = json.loads(reports_data)
                if isinstance(reports_parsed, dict):
                    summaries = []
                    for key, val in reports_parsed.items():
                        if isinstance(val, dict):
                            parts = [f"{k}: {v}" for k, v in val.items() if v]
                            summaries.append(f"{key}: {', '.join(parts)}")
                        else:
                            summaries.append(f"{key}: {val}")
                    reports_data = "; ".join(summaries)
            except Exception:
                pass

        rows.append([
            _stringify(appt.get("id")),                                        # Appointment ID
            _stringify(appt.get("patient_name")),                              # Patient Name
            display_phone,                                                      # Phone
            _stringify(appt.get("patient_age")),                               # Age
            _stringify(appt.get("patient_location") or appt.get("patient_record_location", "")),  # Location
            _stringify(appt.get("date")),                                      # Date
            _stringify(appt.get("time")),                                      # Time
            _stringify(appt.get("reason")),                                    # Reason
            _stringify(appt.get("consultation_type", "")),                     # Consultation Type
            calendar_status,                                                    # Status
            _stringify(appt.get("payment_status", "")),                        # Payment Status
            id_card_display,                                                    # ID Card
            "",                                                                 # ID Card Image URL (local export, no Drive)
            "",                                                                 # Payment Screenshot URL
            "",                                                                 # Reports/Prescriptions URLs
            reports_data,                                                       # Reports Data
            _stringify(appt.get("created_at")),                                # Created At
            _stringify(appt.get("updated_at")),                                # Updated At
            "WhatsApp",                                                         # Source
            _stringify(appt.get("reminder_sent")),                             # Reminder Sent
            _stringify(appt.get("followup_sent")),                             # Followup Sent
            _stringify(appt.get("followup_response")),                         # Followup Response
        ])
    return rows


def export_appointments_xlsx(output_path: Optional[str] = None) -> str:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for Excel export. Run: pip install -r requirements.txt") from exc

    export_dir = Path(settings.EXPORT_DIR)
    export_dir.mkdir(parents=True, exist_ok=True)
    if not output_path:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = str(export_dir / f"appointments_{stamp}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Appointments"
    ws.append(BOOKING_COLUMNS)

    header_fill = PatternFill("solid", fgColor="E8F0FE")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    rows = appointment_rows()
    for row in rows:
        ws.append(row)

    ws.freeze_panes = "A2"
    for idx, column in enumerate(BOOKING_COLUMNS, 1):
        values = [column]
        values.extend(_stringify(row[idx - 1]) for row in rows)
        width = min(max(len(value) for value in values) + 2, 60)
        ws.column_dimensions[get_column_letter(idx)].width = width

    wb.save(output_path)
    return os.path.abspath(output_path)


# ─── Google Drive Upload ─────────────────────────────────────────

_drive_service = None
_drive_creds_source = None
_drive_folder_cache: dict[str, str] = {}

OAUTH_TOKEN_PATH = Path(__file__).parent.parent / "googlekey" / "oauth_token.json"
OAUTH_CREDS_PATH = Path(__file__).parent.parent / "googlekey" / "oauth_credentials.json"


def _get_drive_service(credentials_path: str = ""):
    """Get or create a cached Google Drive service instance.
    Uses OAuth2 token if available, falls back to service account."""
    global _drive_service, _drive_creds_source

    # Prefer OAuth2 token (has user's Drive storage)
    if OAUTH_TOKEN_PATH.exists():
        cache_key = f"oauth:{OAUTH_TOKEN_PATH}"
        if _drive_service and _drive_creds_source == cache_key:
            return _drive_service
        try:
            from google.oauth2.credentials import Credentials
            from googleapiclient.discovery import build

            with open(OAUTH_TOKEN_PATH) as f:
                token_data = json.load(f)

            creds = Credentials(
                token=token_data.get("token"),
                refresh_token=token_data.get("refresh_token"),
                token_uri=token_data.get("token_uri"),
                client_id=token_data.get("client_id"),
                client_secret=token_data.get("client_secret"),
                scopes=token_data.get("scopes"),
            )

            # Auto-refresh if expired
            if creds.expired and creds.refresh_token:
                from google.auth.transport.requests import Request
                creds.refresh(Request())
                token_data["token"] = creds.token
                with open(OAUTH_TOKEN_PATH, "w") as f:
                    json.dump(token_data, f, indent=2)
                logger.info("OAuth token refreshed")

            _drive_service = build('drive', 'v3', credentials=creds)
            _drive_creds_source = cache_key
            logger.info("Drive service initialized with OAuth2 credentials")
            return _drive_service
        except Exception as e:
            logger.warning(f"OAuth2 Drive auth failed: {e}. Falling back to service account.")

    # Fallback to service account (limited - no personal storage)
    if credentials_path:
        cache_key = f"sa:{credentials_path}"
        if _drive_service and _drive_creds_source == cache_key:
            return _drive_service
        try:
            from google.oauth2 import service_account
            from googleapiclient.discovery import build

            SCOPES = ['https://www.googleapis.com/auth/drive']
            creds = service_account.Credentials.from_service_account_file(credentials_path, scopes=SCOPES)
            _drive_service = build('drive', 'v3', credentials=creds)
            _drive_creds_source = cache_key
            logger.warning("Drive service using service account (uploads may fail without Shared Drive)")
            return _drive_service
        except Exception as e:
            logger.error(f"Failed to create Drive service: {e}")
            return None

    return None


def _get_mime_type(file_path: str) -> str:
    """Get MIME type from file extension."""
    ext = Path(file_path).suffix.lower()
    mime_map = {
        '.jpg': 'image/jpeg',
        '.jpeg': 'image/jpeg',
        '.png': 'image/png',
        '.webp': 'image/webp',
        '.gif': 'image/gif',
        '.bmp': 'image/bmp',
        '.tiff': 'image/tiff',
        '.pdf': 'application/pdf',
        '.wav': 'audio/wav',
        '.mp3': 'audio/mpeg',
        '.ogg': 'audio/ogg',
    }
    return mime_map.get(ext, 'application/octet-stream')


def _get_or_create_subfolder(service, parent_id: str, folder_name: str) -> str:
    """Get or create a subfolder in Google Drive. Returns folder ID."""
    cache_key = f"{parent_id}:{folder_name}"
    if cache_key in _drive_folder_cache:
        return _drive_folder_cache[cache_key]
    try:
        safe_name = folder_name.replace("'", "\\'")
        query = f"name='{safe_name}' and '{parent_id}' in parents and mimeType='application/vnd.google-apps.folder' and trashed=false"
        results = service.files().list(q=query, fields="files(id, name)").execute()
        files = results.get("files", [])
        if files:
            folder_id = files[0]["id"]
            _drive_folder_cache[cache_key] = folder_id
            return folder_id
        folder_metadata = {
            "name": folder_name,
            "mimeType": "application/vnd.google-apps.folder",
            "parents": [parent_id],
        }
        folder = service.files().create(body=folder_metadata, fields="id").execute()
        folder_id = folder["id"]
        _drive_folder_cache[cache_key] = folder_id
        logger.info(f"Created Drive subfolder: {folder_name}")
        return folder_id
    except Exception as e:
        logger.error(f"Failed to get/create subfolder '{folder_name}': {e}")
        return parent_id


def upload_file_to_drive(file_path: str, credentials_path: str = "", file_label: str = "Document", patient_name: str = "Patient", patient_phone: str = "") -> str:
    """Upload any file to Google Drive with organized patient subfolders. Returns shareable URL."""
    if not file_path or not os.path.exists(file_path):
        return ""

    service = _get_drive_service(credentials_path)
    if not service:
        return ""

    try:
        from googleapiclient.http import MediaFileUpload

        ext = Path(file_path).suffix
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{file_label}_{timestamp}{ext}"
        mime_type = _get_mime_type(file_path)

        # Determine parent folder
        folder_id = settings.GOOGLE_DRIVE_FOLDER_ID
        if folder_id and patient_name:
            safe_name = patient_name.replace(" ", "_").replace("/", "_").replace("\\", "_")
            folder_label = f"{safe_name}_{patient_phone[-4:]}" if patient_phone else safe_name
            folder_id = _get_or_create_subfolder(service, folder_id, folder_label)

        file_metadata = {'name': filename}
        if folder_id:
            file_metadata['parents'] = [folder_id]

        media = MediaFileUpload(file_path, mimetype=mime_type, resumable=True)
        uploaded = service.files().create(
            body=file_metadata,
            media_body=media,
            fields='id',
        ).execute()

        file_id = uploaded.get('id')
        if not file_id:
            return ""

        service.permissions().create(
            fileId=file_id,
            body={'type': 'anyone', 'role': 'reader'},
        ).execute()

        url = f"https://drive.google.com/file/d/{file_id}/view"
        logger.info(f"Uploaded {file_label} for {patient_name}: {url}")
        return url

    except Exception as e:
        logger.error(f"Failed to upload {file_label} to Drive: {e}")
        return ""


def upload_patient_documents_to_drive(phone: str, credentials_path: str = "", patient_name: str = "Patient") -> dict:
    """Upload all documents for a patient to Drive. Returns dict of URLs by type."""
    docs = get_patient_documents(phone)
    urls = {
        "id_card": "",
        "payment_screenshot": "",
        "prescription": [],
        "report": [],
    }

    for doc in docs:
        doc_type = doc.get("document_type", "general")
        file_path = doc.get("file_path", "")
        if not file_path or not os.path.exists(file_path):
            continue

        label_map = {
            "id_card": "IDCard",
            "payment_screenshot": "Payment",
            "prescription": "Prescription",
            "report": "Report",
        }
        label = label_map.get(doc_type, "Document")
        url = upload_file_to_drive(file_path, credentials_path, label, patient_name, phone)

        if not url:
            continue

        if doc_type == "id_card" and not urls["id_card"]:
            urls["id_card"] = url
        elif doc_type == "payment_screenshot" and not urls["payment_screenshot"]:
            urls["payment_screenshot"] = url
        elif doc_type == "prescription":
            urls["prescription"].append(url)
        elif doc_type == "report":
            urls["report"].append(url)

    return urls


def _get_document_urls_for_appointment(appt: dict, credentials_path: str, doc_urls_cache: dict) -> dict:
    """Get Drive URLs for all documents linked to an appointment."""
    phone = appt.get("phone", "")
    patient_name = appt.get("patient_name", "Patient")

    if phone not in doc_urls_cache:
        doc_urls_cache[phone] = upload_patient_documents_to_drive(phone, credentials_path, patient_name)

    return doc_urls_cache[phone]


# ─── Google Sheet Sync ───────────────────────────────────────────

def sync_appointments_to_google_sheet(
    sheet_id: Optional[str] = None,
    credentials_path: Optional[str] = None,
    worksheet_gid: Optional[int] = None,
) -> dict:
    sheet_id = sheet_id or settings.GOOGLE_SHEET_ID
    credentials_path = credentials_path or settings.GOOGLE_SERVICE_ACCOUNT_JSON

    if not sheet_id:
        raise ValueError("Google Sheet ID is missing.")
    if not credentials_path:
        raise ValueError("Google service account JSON path is missing.")
    if not os.path.exists(credentials_path):
        raise FileNotFoundError(f"Google service account file not found: {credentials_path}")

    try:
        import gspread
    except ImportError as exc:
        raise RuntimeError("gspread and google-auth are required. Run: pip install -r requirements.txt") from exc

    client = gspread.service_account(filename=credentials_path)
    spreadsheet = client.open_by_key(sheet_id)

    gid = worksheet_gid if worksheet_gid is not None else settings.GOOGLE_SHEET_GID
    if gid is not None:
        worksheet = spreadsheet.get_worksheet_by_id(gid)
    else:
        worksheet = spreadsheet.sheet1

    if worksheet is None:
        worksheet = spreadsheet.sheet1

    existing_rows = worksheet.get_all_values()

    # Ensure header row matches our columns
    if not existing_rows:
        worksheet.append_row(SHEET_COLUMNS, value_input_option="USER_ENTERED")
        existing_rows = [SHEET_COLUMNS]
    else:
        # Update header to include any new columns (preserves existing data)
        current_header = existing_rows[0]
        new_cols = [c for c in SHEET_COLUMNS if c not in current_header]
        if new_cols:
            updated_header = current_header + new_cols
            worksheet.update('A1', [updated_header], value_input_option="USER_ENTERED")
            existing_rows[0] = updated_header
            logger.info(f"Added {len(new_cols)} new columns to Sheet header: {new_cols}")

    # Build column index map (handles both old and new column formats)
    header = existing_rows[0] if existing_rows else SHEET_COLUMNS
    col_map = {}
    for col_name in SHEET_COLUMNS:
        if col_name in header:
            col_map[col_name] = header.index(col_name)

    def _get_col(name, default=0):
        return col_map.get(name, default)

    appt_id_col = _get_col("Appointment ID")
    date_col = _get_col("Date")
    time_col = _get_col("Time")

    # Phone column: prefer "Phone Number" (old sheet has data there), fall back to "Phone" (new column)
    phone_col = header.index("Phone Number") if "Phone Number" in header else col_map.get("Phone", 1)

    existing_by_id = {}  # appointment_id -> row_number
    existing_by_key = {}  # (phone, date, time) -> row_number

    def _normalize_phone(p: str) -> str:
        """Strip @lid, @s.whatsapp.net, + prefix for matching."""
        p = p.split('@')[0].strip()
        if p.startswith('+'):
            p = p[1:]
        return p

    for idx, row in enumerate(existing_rows[1:], start=2):  # row 2 = first data row
        if len(row) > appt_id_col and row[appt_id_col].strip():
            existing_by_id[row[appt_id_col].strip()] = idx
        if len(row) > max(phone_col, date_col, time_col):
            key = (_normalize_phone(row[phone_col]), row[date_col], row[time_col])
            existing_by_key[key] = idx

    status_map = {
        "booked": "Created",
        "confirmed": "Confirmed",
        "cancelled": "Cancelled",
        "rescheduled": "Rescheduled",
        "no_show": "No Show",
        "checked_in": "Checked In",
    }

    init_db()
    all_appts = get_all_appointments()
    doc_urls_cache = {}  # phone -> {id_card, payment_screenshot, prescription[], report[]}

    new_rows = []
    update_cells = []

    for appt in all_appts:
        appt_id = str(appt.get("id", ""))
        phone = _stringify(appt.get("phone"))
        date = _stringify(appt.get("date"))
        time = _stringify(appt.get("time"))

        # Get document URLs from Google Drive
        doc_urls = _get_document_urls_for_appointment(appt, credentials_path, doc_urls_cache)

        # Parse details_json for contact number
        contact_number = ""
        details_json = appt.get("details_json")
        if details_json:
            try:
                parsed = json.loads(details_json)
                contact_number = parsed.get("contact_number", "")
            except Exception:
                pass
        display_phone = contact_number if contact_number else phone

        raw_status = _stringify(appt.get("status"))
        calendar_status = status_map.get(raw_status, raw_status)

        # ID card display
        id_card_val = _stringify(appt.get("id_card"))
        id_card_image = appt.get("id_card_image_path")
        if id_card_image:
            id_card_display = "ID image on file"
        elif id_card_val:
            id_card_display = id_card_val
        else:
            id_card_display = ""

        # Reports data (extracted JSON summary)
        reports_data = _stringify(appt.get("reports_data_json", ""))
        if reports_data:
            try:
                reports_parsed = json.loads(reports_data)
                # Make it human-readable
                if isinstance(reports_parsed, dict):
                    summaries = []
                    for key, val in reports_parsed.items():
                        if isinstance(val, dict):
                            parts = [f"{k}: {v}" for k, v in val.items() if v]
                            summaries.append(f"{key}: {', '.join(parts)}")
                        else:
                            summaries.append(f"{key}: {val}")
                    reports_data = "; ".join(summaries)
            except Exception:
                pass

        # URLs
        id_card_url = doc_urls.get("id_card", "")
        payment_url = doc_urls.get("payment_screenshot", "")
        report_urls = doc_urls.get("report", []) + doc_urls.get("prescription", [])
        reports_urls_str = ", ".join(report_urls) if report_urls else ""

        # Build the row matching SHEET_COLUMNS order
        row_data = [
            appt_id,                                        # Appointment ID
            _stringify(appt.get("patient_name")),           # Patient Name
            display_phone,                                  # Phone
            _stringify(appt.get("patient_age")),            # Age
            _stringify(appt.get("patient_location") or appt.get("patient_record_location", "")),  # Location
            date,                                           # Date
            time,                                           # Time
            _stringify(appt.get("reason")),                 # Reason
            _stringify(appt.get("consultation_type", "")),  # Consultation Type
            calendar_status,                                # Status
            _stringify(appt.get("payment_status", "")),     # Payment Status
            id_card_display,                                # ID Card
            id_card_url,                                    # ID Card Image URL
            payment_url,                                    # Payment Screenshot URL
            reports_urls_str,                               # Reports/Prescriptions URLs
            reports_data,                                   # Reports Data
            _stringify(appt.get("created_at")),             # Created At
            _stringify(appt.get("updated_at")),             # Updated At
            "WhatsApp",                                     # Source
            _stringify(appt.get("reminder_sent")),          # Reminder Sent
            _stringify(appt.get("followup_sent")),          # Followup Sent
            _stringify(appt.get("followup_response")),      # Followup Response
        ]

        # Check if row exists (by ID first, then by key)
        target_row = None
        if appt_id and appt_id in existing_by_id:
            target_row = existing_by_id[appt_id]
        elif (_normalize_phone(phone), date, time) in existing_by_key:
            target_row = existing_by_key[(_normalize_phone(phone), date, time)]

        if target_row:
            # Update existing row using actual sheet column positions
            for col_name, value in zip(SHEET_COLUMNS, row_data):
                if col_name in col_map:
                    sheet_col = col_map[col_name] + 1  # gspread is 1-indexed
                    update_cells.append(gspread.Cell(target_row, sheet_col, value))
        else:
            # New row
            new_rows.append(row_data)

    # Batch update existing rows
    if update_cells:
        worksheet.update_cells(update_cells, value_input_option="USER_ENTERED")
        logger.info(f"Updated {len(update_cells)} cells in Google Sheet")

    # Append new rows
    if new_rows:
        worksheet.append_rows(new_rows, value_input_option="USER_ENTERED")
        logger.info(f"Appended {len(new_rows)} new rows to Google Sheet")

    # Count unique updated rows from cell list
    updated_rows = len(set(cell.row for cell in update_cells)) if update_cells else 0

    return {
        "spreadsheet": spreadsheet.title,
        "worksheet": worksheet.title,
        "rows_synced": len(new_rows),
        "rows_updated": updated_rows,
    }


# ─── Google Calendar ─────────────────────────────────────────────

def sync_appointment_to_google_calendar(appointment: dict) -> dict:
    calendar_id = settings.GOOGLE_CALENDAR_ID
    credentials_path = settings.GOOGLE_SERVICE_ACCOUNT_JSON

    if not calendar_id or calendar_id == "primary":
        calendar_id = "primary"

    if not credentials_path or not os.path.exists(credentials_path):
        logger.warning("Google Calendar Sync: Credentials not found.")
        return {}

    try:
        from google.oauth2 import service_account
        from googleapiclient.discovery import build
    except ImportError as exc:
        logger.error("google-api-python-client is required. Run: pip install -r requirements.txt")
        return {}

    SCOPES = ['https://www.googleapis.com/auth/calendar.events']

    try:
        creds = service_account.Credentials.from_service_account_file(
            credentials_path, scopes=SCOPES)
        service = build('calendar', 'v3', credentials=creds)

        date_str = appointment.get("date")
        time_str = appointment.get("time")

        start_dt = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M")
        end_dt = start_dt + timedelta(minutes=settings.SLOT_DURATION_MINUTES)

        timezone = settings.CLINIC_TIMEZONE

        name = appointment.get("patient_name", "Patient")
        reason = appointment.get("reason", "No reason provided")

        contact_number = ""
        details_json = appointment.get("details_json")
        if details_json:
            try:
                parsed = json.loads(details_json)
                contact_number = parsed.get("contact_number", "")
            except Exception:
                pass

        display_phone = contact_number if contact_number else appointment.get("phone", "")

        event = {
            'summary': f'Appointment: {name}',
            'description': f'Phone: {display_phone}\nReason: {reason}',
            'start': {
                'dateTime': start_dt.isoformat(),
                'timeZone': timezone,
            },
            'end': {
                'dateTime': end_dt.isoformat(),
                'timeZone': timezone,
            },
        }

        event_result = service.events().insert(calendarId=calendar_id, body=event).execute()

        logger.info(f"Successfully synced appointment {appointment.get('id')} to Google Calendar.")
        return {
            "status": "success",
            "event_link": event_result.get('htmlLink'),
            "event_id": event_result.get('id')
        }

    except Exception as e:
        logger.error(f"Failed to sync appointment to Google Calendar: {e}")
        return {"status": "error", "message": str(e)}
